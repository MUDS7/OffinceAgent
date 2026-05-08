from __future__ import annotations

from base64 import b64encode
from io import BytesIO
from pathlib import Path
from typing import Any, Literal

import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from pydantic import BaseModel, Field


BasicExcelCommand = Literal["set_cell", "set_range", "insert_row", "insert_column", "delete_row", "split_column"]
AdvancedExcelCommand = Literal["fill_empty_cells", "summarize_by_column", "generate_report"]
ExcelCommandName = BasicExcelCommand | AdvancedExcelCommand
ExcelCommandCategory = Literal["basic", "advanced"]

BASIC_EXCEL_COMMANDS: tuple[BasicExcelCommand, ...] = (
    "set_cell",
    "set_range",
    "insert_row",
    "insert_column",
    "delete_row",
    "split_column",
)
ADVANCED_EXCEL_COMMANDS: tuple[AdvancedExcelCommand, ...] = (
    "fill_empty_cells",
    "summarize_by_column",
    "generate_report",
)


class ExcelCommandSpec(BaseModel):
    command: ExcelCommandName
    category: ExcelCommandCategory
    description: str
    required_args: list[str] = Field(default_factory=list)
    optional_args: list[str] = Field(default_factory=list)


class ExcelCommandsResponse(BaseModel):
    basic: list[ExcelCommandSpec]
    advanced: list[ExcelCommandSpec]


class ExcelExecuteRequest(BaseModel):
    command: ExcelCommandName
    file_path: str | None = None
    path: str | None = None
    sheet: str | None = None
    output_path: str | None = None
    save_to_disk: bool = False
    args: dict[str, Any] = Field(default_factory=dict)


class ExcelExecuteResponse(BaseModel):
    command: ExcelCommandName
    category: ExcelCommandCategory
    file_path: str
    output_path: str
    workbook_base64: str | None = None
    saved_to_disk: bool = False
    sheet: str | None = None
    rows_affected: int = 0
    cells_affected: int = 0
    summary: str
    data: list[dict[str, Any]] | None = None


def get_excel_commands() -> ExcelCommandsResponse:
    return ExcelCommandsResponse(
        basic=[
            ExcelCommandSpec(
                command="set_cell",
                category="basic",
                description="Set one worksheet cell to a value.",
                required_args=["cell", "value"],
            ),
            ExcelCommandSpec(
                command="set_range",
                category="basic",
                description="Set a rectangular worksheet range from a two-dimensional values array.",
                required_args=["values"],
                optional_args=["start_cell", "range"],
            ),
            ExcelCommandSpec(
                command="insert_row",
                category="basic",
                description="Insert one or more rows by index, before/after a row or range, or in the middle of the used sheet.",
                optional_args=["index", "before_row", "after_row", "range", "position", "amount", "values"],
            ),
            ExcelCommandSpec(
                command="insert_column",
                category="basic",
                description="Insert one or more columns by index, letter, before/after a column or range, or in the middle of the used sheet.",
                optional_args=[
                    "index",
                    "column",
                    "before_column",
                    "after_column",
                    "range",
                    "position",
                    "amount",
                    "values",
                ],
            ),
            ExcelCommandSpec(
                command="delete_row",
                category="basic",
                description="Delete one or more rows from a worksheet.",
                required_args=["index"],
                optional_args=["amount"],
            ),
            ExcelCommandSpec(
                command="split_column",
                category="basic",
                description="Split cell text from one column or selected range into adjacent columns by delimiter, character, fixed positions, or row-specific fixed positions.",
                optional_args=[
                    "range",
                    "source_column",
                    "column",
                    "source_cell",
                    "start_row",
                    "end_row",
                    "target_cell",
                    "target_column",
                    "delimiter",
                    "mode",
                    "positions",
                    "widths",
                    "split_at",
                    "positions_by_row",
                    "widths_by_row",
                    "row_positions",
                    "row_widths",
                    "max_splits",
                    "trim",
                    "insert_columns",
                ],
            ),
        ],
        advanced=[
            ExcelCommandSpec(
                command="fill_empty_cells",
                category="advanced",
                description="Fill blank cells in selected columns or the used range.",
                optional_args=["columns", "fill_value", "method"],
            ),
            ExcelCommandSpec(
                command="summarize_by_column",
                category="advanced",
                description="Group rows by one column and aggregate numeric columns.",
                required_args=["group_by"],
                optional_args=["aggregations", "output_sheet"],
            ),
            ExcelCommandSpec(
                command="generate_report",
                category="advanced",
                description="Generate a formatted summary workbook from an Excel sheet.",
                required_args=["group_by"],
                optional_args=["aggregations", "report_title"],
            ),
        ],
    )


def execute_excel_command(request: ExcelExecuteRequest) -> ExcelExecuteResponse:
    source_path = _resolve_source_path(request)
    output_path = _resolve_output_path(source_path, request.output_path, request.save_to_disk)
    _ensure_xlsx(source_path)

    if request.command == "set_cell":
        return _set_cell(request, source_path, output_path)
    if request.command == "set_range":
        return _set_range(request, source_path, output_path)
    if request.command == "insert_row":
        return _insert_row(request, source_path, output_path)
    if request.command == "insert_column":
        return _insert_column(request, source_path, output_path)
    if request.command == "delete_row":
        return _delete_row(request, source_path, output_path)
    if request.command == "split_column":
        return _split_column(request, source_path, output_path)
    if request.command == "fill_empty_cells":
        return _fill_empty_cells(request, source_path, output_path)
    if request.command == "summarize_by_column":
        return _summarize_by_column(request, source_path, output_path)
    if request.command == "generate_report":
        return _generate_report(request, source_path, output_path)

    raise HTTPException(status_code=400, detail=f"Unsupported Excel command: {request.command}")


def _set_cell(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    cell = str(_required_arg(request.args, "cell")).upper()
    value = request.args.get("value")
    workbook = load_workbook(source_path)
    worksheet = _select_sheet(workbook, request.sheet)
    worksheet[cell] = value

    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        worksheet.title,
        cells_affected=1,
        summary=f"Set {worksheet.title}!{cell}.",
    )


def _set_range(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    values = _normalize_matrix(_required_arg(request.args, "values"))
    start_row, start_column = _range_start(request.args)
    workbook = load_workbook(source_path)
    worksheet = _select_sheet(workbook, request.sheet)

    cells_affected = 0
    for row_offset, row_values in enumerate(values):
        for column_offset, value in enumerate(row_values):
            worksheet.cell(row=start_row + row_offset, column=start_column + column_offset, value=value)
            cells_affected += 1

    end_cell = f"{get_column_letter(start_column + max(len(row) for row in values) - 1)}{start_row + len(values) - 1}"

    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        worksheet.title,
        cells_affected=cells_affected,
        summary=f"Set range {worksheet.title}!{get_column_letter(start_column)}{start_row}:{end_cell}.",
    )


def _insert_row(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    amount = _positive_int(request.args.get("amount", 1), "amount")
    values = request.args.get("values")
    workbook = load_workbook(source_path)
    worksheet = _select_sheet(workbook, request.sheet)
    index = _insert_row_index(request.args, worksheet)
    worksheet.insert_rows(index, amount)

    cells_affected = 0
    if values is not None:
        matrix = _normalize_matrix(values)
        for row_offset, row_values in enumerate(matrix[:amount]):
            for column_offset, value in enumerate(row_values):
                worksheet.cell(row=index + row_offset, column=1 + column_offset, value=value)
                cells_affected += 1

    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        worksheet.title,
        rows_affected=amount,
        cells_affected=cells_affected,
        summary=f"Inserted {amount} row(s) at {worksheet.title}!{index}.",
    )


def _insert_column(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    amount = _positive_int(request.args.get("amount", 1), "amount")
    values = request.args.get("values")
    workbook = load_workbook(source_path)
    worksheet = _select_sheet(workbook, request.sheet)
    index = _insert_column_index(request.args, worksheet)
    worksheet.insert_cols(index, amount)

    cells_affected = 0
    if values is not None:
        matrix = _normalize_column_values(values)
        for row_offset, row_values in enumerate(matrix):
            for column_offset, value in enumerate(row_values[:amount]):
                worksheet.cell(row=1 + row_offset, column=index + column_offset, value=value)
                cells_affected += 1

    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        worksheet.title,
        rows_affected=amount,
        cells_affected=cells_affected,
        summary=f"Inserted {amount} column(s) at {worksheet.title}!{get_column_letter(index)}.",
    )


def _insert_row_index(args: dict[str, Any], worksheet: Any) -> int:
    position = _normalize_insert_position(args.get("position") or args.get("location") or args.get("where"))
    raw_index = args.get("index")
    if raw_index is not None and _normalize_insert_position(raw_index) == "middle":
        return _middle_row_index(worksheet)

    exact_row = _first_positive_int_arg(args, ("index", "insert_at", "row_index", "target_row"))
    if exact_row is not None:
        return exact_row + 1 if position == "after" else exact_row

    before_row = _first_positive_int_arg(args, ("before_row", "above_row"))
    if before_row is not None:
        return before_row

    after_row = _first_positive_int_arg(args, ("after_row", "below_row"))
    if after_row is not None:
        return after_row + 1

    range_address = args.get("range") or args.get("range_address") or args.get("selection")
    range_rows = _range_row_bounds(range_address)
    if range_rows is not None:
        min_row, max_row = range_rows
        if position == "after":
            return max_row + 1
        if position == "middle":
            return min_row + ((max_row - min_row + 1) // 2)
        return min_row

    if position == "middle":
        return _middle_row_index(worksheet)

    raise HTTPException(
        status_code=400,
        detail="args.index, args.before_row, args.after_row, args.range, or args.position='middle' is required",
    )


def _insert_column_index(args: dict[str, Any], worksheet: Any) -> int:
    position = _normalize_insert_position(args.get("position") or args.get("location") or args.get("where"))
    raw_index = args.get("index")
    if raw_index is not None and _normalize_insert_position(raw_index) == "middle":
        return _middle_column_index(worksheet)

    exact_column = _first_column_index_arg(args, ("index", "column", "insert_at", "column_index", "target_column"))
    if exact_column is not None:
        return exact_column + 1 if position == "after" else exact_column

    before_column = _first_column_index_arg(args, ("before_column", "left_of_column"))
    if before_column is not None:
        return before_column

    after_column = _first_column_index_arg(args, ("after_column", "right_of_column"))
    if after_column is not None:
        return after_column + 1

    range_address = args.get("range") or args.get("range_address") or args.get("selection")
    range_columns = _range_column_bounds(range_address)
    if range_columns is not None:
        min_column, max_column = range_columns
        if position == "after":
            return max_column + 1
        if position == "middle":
            return min_column + ((max_column - min_column + 1) // 2)
        return min_column

    if position == "middle":
        return _middle_column_index(worksheet)

    raise HTTPException(
        status_code=400,
        detail="args.index, args.column, args.before_column, args.after_column, args.range, or args.position='middle' is required",
    )


def _first_positive_int_arg(args: dict[str, Any], names: tuple[str, ...]) -> int | None:
    for name in names:
        if name in args and args[name] is not None:
            return _positive_int(args[name], name)
    return None


def _first_column_index_arg(args: dict[str, Any], names: tuple[str, ...]) -> int | None:
    for name in names:
        if name in args and args[name] is not None:
            return _column_index(args[name], name)
    return None


def _normalize_insert_position(value: Any) -> Literal["before", "after", "middle"] | None:
    if value is None:
        return None

    position = str(value).strip().lower()
    if position in ("before", "above", "top", "start", "front", "前", "之前", "上方"):
        return "before"
    if position in ("after", "below", "bottom", "end", "后", "之后", "下方"):
        return "after"
    if position in ("middle", "center", "centre", "between", "中间", "居中"):
        return "middle"

    return None


def _column_index(value: Any, name: str) -> int:
    if isinstance(value, int):
        return _positive_int(value, name)

    raw_value = str(value).strip()
    if "!" in raw_value:
        raw_value = raw_value.rsplit("!", 1)[1]
    raw_value = raw_value.replace("$", "").strip("'")

    if raw_value.isdigit():
        return _positive_int(raw_value, name)

    try:
        min_column, _, _, _ = range_boundaries(raw_value)
        return min_column
    except ValueError:
        pass

    try:
        return column_index_from_string(raw_value.upper())
    except ValueError:
        raise HTTPException(status_code=400, detail=f"args.{name} must be a column letter or index") from None


def _range_row_bounds(value: Any) -> tuple[int, int] | None:
    if value in (None, ""):
        return None

    range_address = str(value).strip()
    if "!" in range_address:
        range_address = range_address.rsplit("!", 1)[1]
    range_address = range_address.replace("$", "").strip("'")

    row_range_parts = range_address.split(":")
    if row_range_parts and all(part.isdigit() for part in row_range_parts):
        row_indexes = [_positive_int(part, "range") for part in row_range_parts]
        return min(row_indexes), max(row_indexes)

    try:
        _, min_row, _, max_row = range_boundaries(range_address)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid range: {value}") from None

    return min_row, max_row


def _range_column_bounds(value: Any) -> tuple[int, int] | None:
    if value in (None, ""):
        return None

    range_address = str(value).strip()
    if "!" in range_address:
        range_address = range_address.rsplit("!", 1)[1]
    range_address = range_address.replace("$", "").strip("'")

    column_range_parts = range_address.split(":")
    if column_range_parts and all(part.isalpha() for part in column_range_parts):
        column_indexes = [_column_index(part, "range") for part in column_range_parts]
        return min(column_indexes), max(column_indexes)

    try:
        min_column, _, max_column, _ = range_boundaries(range_address)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid range: {value}") from None

    return min_column, max_column


def _middle_row_index(worksheet: Any) -> int:
    min_row, max_row = _used_row_bounds(worksheet)
    return min_row + ((max_row - min_row + 1) // 2)


def _used_row_bounds(worksheet: Any) -> tuple[int, int]:
    range_address = worksheet.calculate_dimension()
    try:
        _, min_row, _, max_row = range_boundaries(range_address)
    except ValueError:
        return 1, max(worksheet.max_row, 1)

    return min_row, max(max_row, min_row)


def _middle_column_index(worksheet: Any) -> int:
    min_column, max_column = _used_column_bounds(worksheet)
    return min_column + ((max_column - min_column + 1) // 2)


def _used_column_bounds(worksheet: Any) -> tuple[int, int]:
    range_address = worksheet.calculate_dimension()
    try:
        min_column, _, max_column, _ = range_boundaries(range_address)
    except ValueError:
        return 1, max(worksheet.max_column, 1)

    return min_column, max(max_column, min_column)


def _delete_row(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    index = _positive_int(_required_arg(request.args, "index"), "index")
    amount = _positive_int(request.args.get("amount", 1), "amount")
    workbook = load_workbook(source_path)
    worksheet = _select_sheet(workbook, request.sheet)
    worksheet.delete_rows(index, amount)

    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        worksheet.title,
        rows_affected=amount,
        summary=f"Deleted {amount} row(s) at {worksheet.title}!{index}.",
    )


def _split_column(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    workbook = load_workbook(source_path)
    worksheet = _select_sheet(workbook, request.sheet)
    source_column, start_row, end_row = _split_source_bounds(request.args, worksheet)
    target_row, target_column = _split_target_start(request.args, start_row, source_column)
    split_mode = str(request.args.get("mode", "delimiter")).strip().lower()
    delimiter = request.args.get("delimiter")
    fixed_split_positions = _fixed_split_positions(request.args, split_mode)
    row_split_positions = _row_fixed_split_positions(request.args, start_row, end_row)
    max_splits = _optional_non_negative_int(request.args.get("max_splits"), "max_splits")
    trim_parts = _truthy(request.args.get("trim"))

    split_rows: list[list[Any]] = []
    for row_index in range(start_row, end_row + 1):
        source_value = worksheet.cell(row=row_index, column=source_column).value
        row_fixed_split_positions = row_split_positions.get(row_index, fixed_split_positions)
        split_rows.append(
            _split_cell_value(
                source_value,
                split_mode,
                delimiter,
                max_splits,
                trim_parts,
                row_fixed_split_positions,
            )
        )

    max_width = max((len(row) for row in split_rows), default=0)
    if max_width == 0:
        raise HTTPException(status_code=400, detail="No split output was produced")

    if _truthy(request.args.get("insert_columns")):
        if target_column == source_column:
            if max_width > 1:
                worksheet.insert_cols(source_column + 1, max_width - 1)
        else:
            worksheet.insert_cols(target_column, max_width)

    cells_affected = 0
    for row_offset, row_values in enumerate(split_rows):
        for column_offset in range(max_width):
            value = row_values[column_offset] if column_offset < len(row_values) else None
            worksheet.cell(row=target_row + row_offset, column=target_column + column_offset, value=value)
            cells_affected += 1

    end_cell = f"{get_column_letter(target_column + max_width - 1)}{target_row + len(split_rows) - 1}"
    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        worksheet.title,
        rows_affected=len(split_rows),
        cells_affected=cells_affected,
        summary=(
            f"Split {len(split_rows)} cell(s) from {worksheet.title}!{get_column_letter(source_column)}"
            f" into {worksheet.title}!{get_column_letter(target_column)}{target_row}:{end_cell}."
        ),
    )


def _fill_empty_cells(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    workbook = load_workbook(source_path)
    worksheet = _select_sheet(workbook, request.sheet)
    method = str(request.args.get("method", "value")).lower()
    fill_value = request.args.get("fill_value", "")
    column_indexes = _selected_column_indexes(worksheet, request.args.get("columns"))
    cells_affected = 0

    for column_index in column_indexes:
        previous_value: Any = None
        pending_blank_cells = []
        for row_index in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.value in (None, ""):
                if method == "value":
                    cell.value = fill_value
                    cells_affected += 1
                elif method == "forward":
                    if previous_value not in (None, ""):
                        cell.value = previous_value
                        cells_affected += 1
                elif method == "backward":
                    pending_blank_cells.append(cell)
                else:
                    raise HTTPException(status_code=400, detail="method must be value, forward, or backward")
            else:
                if method == "backward":
                    for blank_cell in pending_blank_cells:
                        blank_cell.value = cell.value
                        cells_affected += 1
                    pending_blank_cells = []
                previous_value = cell.value

    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        worksheet.title,
        cells_affected=cells_affected,
        summary=f"Filled {cells_affected} empty cell(s) in {worksheet.title}.",
    )


def _summarize_by_column(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    group_by = str(_required_arg(request.args, "group_by"))
    output_sheet = str(request.args.get("output_sheet", "Summary"))
    dataframe = _read_sheet_dataframe(source_path, request.sheet)
    summary = _build_summary_dataframe(dataframe, group_by, request.args.get("aggregations"))

    workbook = load_workbook(source_path)
    if output_sheet in workbook.sheetnames:
        del workbook[output_sheet]
    worksheet = workbook.create_sheet(output_sheet)
    _write_dataframe_to_worksheet(summary, worksheet)

    return _finish_openpyxl_response(
        request,
        source_path,
        output_path,
        workbook,
        output_sheet,
        rows_affected=len(summary),
        summary=f"Created summary sheet '{output_sheet}' grouped by '{group_by}'.",
        data=_json_records(summary),
    )


def _generate_report(request: ExcelExecuteRequest, source_path: Path, output_path: Path) -> ExcelExecuteResponse:
    if request.output_path is None:
        output_path = source_path.with_name(f"{source_path.stem}_report.xlsx")

    group_by = str(_required_arg(request.args, "group_by"))
    report_title = str(request.args.get("report_title", "Excel Summary Report"))
    dataframe = _read_sheet_dataframe(source_path, request.sheet)
    summary = _build_summary_dataframe(dataframe, group_by, request.args.get("aggregations"))

    workbook_base64: str | None = None
    writer_target: Path | BytesIO
    if request.save_to_disk:
        writer_target = output_path
    else:
        writer_target = BytesIO()

    with pd.ExcelWriter(writer_target, engine="xlsxwriter") as writer:
        summary.to_excel(writer, sheet_name="Summary", startrow=2, index=False)
        workbook = writer.book
        worksheet = writer.sheets["Summary"]
        title_format = workbook.add_format({"bold": True, "font_size": 16})
        header_format = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "border": 1})
        number_format = workbook.add_format({"num_format": "#,##0.00"})

        worksheet.write(0, 0, report_title, title_format)
        for column_index, column_name in enumerate(summary.columns):
            worksheet.write(2, column_index, column_name, header_format)
            width = max(12, min(36, len(str(column_name)) + 4))
            worksheet.set_column(column_index, column_index, width, number_format if column_index > 0 else None)
        worksheet.freeze_panes(3, 0)
        worksheet.autofilter(2, 0, 2 + len(summary), max(len(summary.columns) - 1, 0))

    if not request.save_to_disk:
        workbook_base64 = _encode_bytes(writer_target.getvalue())

    return _response(
        request,
        source_path,
        output_path,
        "Summary",
        rows_affected=len(summary),
        summary=f"Generated report workbook grouped by '{group_by}'.",
        data=_json_records(summary),
        workbook_base64=workbook_base64,
        saved_to_disk=request.save_to_disk,
    )


def _resolve_source_path(request: ExcelExecuteRequest) -> Path:
    raw_path = request.file_path or request.path
    if not raw_path:
        raise HTTPException(status_code=400, detail="file_path is required")

    path = Path(raw_path).expanduser().resolve()
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Excel file not found: {path}")
    if not path.is_file():
        raise HTTPException(status_code=400, detail=f"Excel path is not a file: {path}")
    return path


def _resolve_output_path(source_path: Path, output_path: str | None, save_to_disk: bool) -> Path:
    if not output_path:
        return source_path

    resolved = Path(output_path).expanduser().resolve()
    if save_to_disk:
        resolved.parent.mkdir(parents=True, exist_ok=True)
    return resolved


def _ensure_xlsx(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported for Excel commands")


def _select_sheet(workbook: Any, sheet_name: str | None) -> Any:
    if not sheet_name:
        return workbook.active
    if sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=404, detail=f"Worksheet not found: {sheet_name}")
    return workbook[sheet_name]


def _required_arg(args: dict[str, Any], name: str) -> Any:
    if name not in args or args[name] is None:
        raise HTTPException(status_code=400, detail=f"args.{name} is required")
    return args[name]


def _positive_int(value: Any, name: str) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"args.{name} must be an integer") from None
    if number < 1:
        raise HTTPException(status_code=400, detail=f"args.{name} must be greater than 0")
    return number


def _normalize_matrix(values: Any) -> list[list[Any]]:
    if not isinstance(values, list) or len(values) == 0:
        raise HTTPException(status_code=400, detail="args.values must be a non-empty list")
    if all(not isinstance(item, list) for item in values):
        return [values]
    if not all(isinstance(item, list) for item in values):
        raise HTTPException(status_code=400, detail="args.values must be a one- or two-dimensional list")
    if any(len(row) == 0 for row in values):
        raise HTTPException(status_code=400, detail="args.values rows cannot be empty")
    return values


def _normalize_column_values(values: Any) -> list[list[Any]]:
    if isinstance(values, list) and values and all(not isinstance(item, list) for item in values):
        return [[item] for item in values]

    return _normalize_matrix(values)


def _split_source_bounds(args: dict[str, Any], worksheet: Any) -> tuple[int, int, int]:
    range_address = args.get("range") or args.get("range_address") or args.get("selection")
    if range_address:
        min_column, min_row, max_column, max_row = _range_bounds(range_address)
        if min_column != max_column:
            raise HTTPException(status_code=400, detail="split_column requires a single source column range")
        return min_column, min_row, max_row

    source_cell = args.get("source_cell") or args.get("cell")
    if source_cell:
        min_column, min_row, _, _ = _range_bounds(source_cell)
        return min_column, min_row, min_row

    source_column = _first_column_index_arg(args, ("source_column", "column"))
    if source_column is not None:
        used_start_row, used_end_row = _used_row_bounds(worksheet)
        start_row = _positive_int(args.get("start_row", used_start_row), "start_row")
        end_row = _positive_int(args.get("end_row", used_end_row), "end_row")
        if end_row < start_row:
            raise HTTPException(status_code=400, detail="args.end_row must be greater than or equal to args.start_row")
        return source_column, start_row, end_row

    raise HTTPException(status_code=400, detail="args.range, args.source_cell, or args.source_column is required")


def _split_target_start(args: dict[str, Any], start_row: int, source_column: int) -> tuple[int, int]:
    target_cell = args.get("target_cell") or args.get("start_cell")
    if target_cell:
        target_column, target_row, _, _ = _range_bounds(target_cell)
        return target_row, target_column

    target_column = _first_column_index_arg(args, ("target_column", "output_column"))
    target_row = _positive_int(args.get("target_row", start_row), "target_row")
    return target_row, target_column or source_column


def _split_cell_value(
    value: Any,
    split_mode: str,
    delimiter: Any,
    max_splits: int | None,
    trim_parts: bool,
    fixed_split_positions: list[int] | None = None,
) -> list[Any]:
    text = "" if value is None else str(value)
    if split_mode in ("character", "characters", "char", "chars", "each_character", "每字符", "按字符", "字符"):
        parts = list(text) if text else [""]
        if max_splits is not None and len(parts) > max_splits + 1:
            parts = parts[:max_splits] + ["".join(parts[max_splits:])]
    elif _is_fixed_split_mode(split_mode):
        if fixed_split_positions is None:
            raise HTTPException(
                status_code=400,
                detail="args.positions, args.split_at, or args.widths is required for fixed position split mode",
            )
        parts = _split_text_at_positions(text, fixed_split_positions)
    else:
        normalized_delimiter = _normalize_delimiter(delimiter)
        if normalized_delimiter == "":
            raise HTTPException(status_code=400, detail="args.delimiter cannot be empty; use mode='characters' instead")
        parts = text.split(normalized_delimiter, max_splits if max_splits is not None else -1)

    if trim_parts:
        return [part.strip() if isinstance(part, str) else part for part in parts]
    return parts


def _is_fixed_split_mode(split_mode: str) -> bool:
    return split_mode in (
        "fixed",
        "fixed_width",
        "fixed_widths",
        "fixed_position",
        "fixed_positions",
        "position",
        "positions",
        "slice",
        "slices",
    )


def _fixed_split_positions(args: dict[str, Any], split_mode: str) -> list[int] | None:
    raw_positions = _first_present_arg(
        args,
        (
            "positions",
            "split_positions",
            "fixed_positions",
            "split_at",
            "cut_positions",
        ),
    )
    if raw_positions not in (None, ""):
        return _normalize_positive_int_list(raw_positions, "positions")

    raw_widths = _first_present_arg(args, ("widths", "fixed_widths", "column_widths"))
    if raw_widths not in (None, ""):
        widths = _normalize_positive_int_list(raw_widths, "widths", require_increasing=False)
        return _positions_from_widths(widths)

    if _is_fixed_split_mode(split_mode):
        return None
    return None


def _row_fixed_split_positions(args: dict[str, Any], start_row: int, end_row: int) -> dict[int, list[int]]:
    raw_positions = _first_present_arg(
        args,
        (
            "positions_by_row",
            "split_positions_by_row",
            "fixed_positions_by_row",
            "row_positions",
        ),
    )
    if raw_positions not in (None, ""):
        return _normalize_row_split_positions(raw_positions, "positions_by_row", start_row, end_row)

    raw_widths = _first_present_arg(
        args,
        (
            "widths_by_row",
            "fixed_widths_by_row",
            "row_widths",
        ),
    )
    if raw_widths not in (None, ""):
        return _normalize_row_split_positions(raw_widths, "widths_by_row", start_row, end_row, value_is_widths=True)

    return {}


def _normalize_row_split_positions(
    value: Any,
    name: str,
    start_row: int,
    end_row: int,
    *,
    value_is_widths: bool = False,
) -> dict[int, list[int]]:
    row_count = end_row - start_row + 1
    positions_by_row: dict[int, list[int]] = {}

    if isinstance(value, dict):
        for raw_row, raw_positions in value.items():
            row_index = _row_index_key(raw_row, name)
            _ensure_row_in_split_range(row_index, start_row, end_row, name)
            positions_by_row[row_index] = _normalize_position_or_widths_value(
                raw_positions,
                f"{name}.{raw_row}",
                value_is_widths=value_is_widths,
            )
        return positions_by_row

    if isinstance(value, list):
        object_rows = [item for item in value if isinstance(item, dict) and "row" in item]
        if object_rows:
            if len(object_rows) != len(value):
                raise HTTPException(
                    status_code=400,
                    detail=f"args.{name} row-object form cannot be mixed with aligned list items",
                )
            for item in object_rows:
                row_index = _positive_int(item.get("row"), f"{name}[].row")
                _ensure_row_in_split_range(row_index, start_row, end_row, name)
                raw_positions = _row_split_item_value(item, value_is_widths)
                positions_by_row[row_index] = _normalize_position_or_widths_value(
                    raw_positions,
                    f"{name}[]",
                    value_is_widths=value_is_widths,
                )
            return positions_by_row

        if len(value) != row_count:
            raise HTTPException(
                status_code=400,
                detail=f"args.{name} must contain exactly {row_count} item(s) for the selected source rows",
            )
        for offset, raw_positions in enumerate(value):
            if raw_positions in (None, ""):
                continue
            positions_by_row[start_row + offset] = _normalize_position_or_widths_value(
                raw_positions,
                f"{name}[{offset}]",
                value_is_widths=value_is_widths,
            )
        return positions_by_row

    raise HTTPException(status_code=400, detail=f"args.{name} must be a list or an object keyed by row number")


def _row_index_key(value: Any, name: str) -> int:
    if isinstance(value, int):
        return _positive_int(value, name)

    text = str(value).strip().lower()
    if text.startswith("row_"):
        text = text[4:]
    elif text.startswith("row"):
        text = text[3:].strip("_ ")
    return _positive_int(text, name)


def _ensure_row_in_split_range(row_index: int, start_row: int, end_row: int, name: str) -> None:
    if row_index < start_row or row_index > end_row:
        raise HTTPException(
            status_code=400,
            detail=f"args.{name} row {row_index} is outside the source range {start_row}:{end_row}",
        )


def _row_split_item_value(item: dict[str, Any], value_is_widths: bool) -> Any:
    if value_is_widths:
        return _first_present_arg(item, ("widths", "fixed_widths", "value"))
    return _first_present_arg(item, ("positions", "split_at", "value"))


def _normalize_position_or_widths_value(value: Any, name: str, *, value_is_widths: bool) -> list[int]:
    if value_is_widths:
        widths = _normalize_positive_int_list(value, name, require_increasing=False)
        return _positions_from_widths(widths)
    return _normalize_positive_int_list(value, name)


def _first_present_arg(args: dict[str, Any], names: tuple[str, ...]) -> Any:
    for name in names:
        if name in args:
            return args[name]
    return None


def _normalize_positive_int_list(value: Any, name: str, require_increasing: bool = True) -> list[int]:
    if isinstance(value, int):
        raw_items = [value]
    elif isinstance(value, str):
        raw_items = [item.strip() for item in value.replace(";", ",").split(",") if item.strip()]
    elif isinstance(value, list):
        raw_items = value
    else:
        raise HTTPException(status_code=400, detail=f"args.{name} must be an integer or a list of integers")

    if not raw_items:
        raise HTTPException(status_code=400, detail=f"args.{name} cannot be empty")

    numbers: list[int] = []
    for index, item in enumerate(raw_items):
        numbers.append(_positive_int(item, f"{name}[{index}]"))

    if require_increasing:
        previous = 0
        for number in numbers:
            if number <= previous:
                raise HTTPException(status_code=400, detail=f"args.{name} must be strictly increasing")
            previous = number
    return numbers


def _positions_from_widths(widths: list[int]) -> list[int]:
    positions: list[int] = []
    current_position = 0
    for width in widths:
        current_position += width
        positions.append(current_position)
    return positions


def _split_text_at_positions(text: str, positions: list[int]) -> list[str]:
    parts: list[str] = []
    start = 0
    for position in positions:
        parts.append(text[start:position])
        start = position
    parts.append(text[start:])
    return parts or [""]


def _normalize_delimiter(value: Any) -> str:
    if value is None:
        raise HTTPException(status_code=400, detail="args.delimiter is required for delimiter split mode")

    delimiter = str(value)
    aliases = {
        "\\t": "\t",
        "\\n": "\n",
        "tab": "\t",
        "tabs": "\t",
        "制表符": "\t",
        "space": " ",
        "spaces": " ",
        "whitespace": " ",
        "空格": " ",
        "comma": ",",
        "逗号": ",",
        "semicolon": ";",
        "分号": ";",
    }
    return aliases.get(delimiter.strip().lower(), delimiter)


def _optional_non_negative_int(value: Any, name: str) -> int | None:
    if value is None:
        return None
    try:
        number = int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"args.{name} must be an integer") from None
    if number < 0:
        raise HTTPException(status_code=400, detail=f"args.{name} must be greater than or equal to 0")
    return number


def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in ("1", "true", "yes", "y", "on", "是")


def _range_bounds(value: Any) -> tuple[int, int, int, int]:
    range_address = str(value).strip()
    if "!" in range_address:
        range_address = range_address.rsplit("!", 1)[1]
    range_address = range_address.replace("$", "").strip("'")
    try:
        return range_boundaries(range_address)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid range: {value}") from None


def _range_start(args: dict[str, Any]) -> tuple[int, int]:
    range_address = args.get("range") or args.get("range_address")
    if range_address:
        min_col, min_row, _, _ = range_boundaries(str(range_address))
        return min_row, min_col

    start_cell = str(args.get("start_cell") or args.get("cell") or "A1")
    min_col, min_row, _, _ = range_boundaries(start_cell)
    return min_row, min_col


def _selected_column_indexes(worksheet: Any, columns: Any) -> list[int]:
    if columns in (None, []):
        return list(range(1, worksheet.max_column + 1))
    if not isinstance(columns, list):
        raise HTTPException(status_code=400, detail="args.columns must be a list")

    header_to_index = {
        str(worksheet.cell(row=1, column=column_index).value): column_index
        for column_index in range(1, worksheet.max_column + 1)
        if worksheet.cell(row=1, column=column_index).value not in (None, "")
    }
    indexes: list[int] = []
    for column in columns:
        if isinstance(column, int):
            indexes.append(_positive_int(column, "columns[]"))
        else:
            label = str(column)
            if label in header_to_index:
                indexes.append(header_to_index[label])
            else:
                try:
                    indexes.append(column_index_from_string(label.upper()))
                except ValueError:
                    raise HTTPException(status_code=400, detail=f"Unknown column: {label}") from None
    return indexes


def _read_sheet_dataframe(source_path: Path, sheet_name: str | None) -> pd.DataFrame:
    try:
        return pd.read_excel(source_path, sheet_name=sheet_name or 0)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from None


def _build_summary_dataframe(
    dataframe: pd.DataFrame,
    group_by: str,
    aggregations: Any,
) -> pd.DataFrame:
    if group_by not in dataframe.columns:
        raise HTTPException(status_code=400, detail=f"group_by column not found: {group_by}")

    if aggregations is None:
        numeric_columns = [column for column in dataframe.select_dtypes(include="number").columns if column != group_by]
        if not numeric_columns:
            summary = dataframe.groupby(group_by, dropna=False).size().reset_index(name="count")
        else:
            summary = dataframe.groupby(group_by, dropna=False)[numeric_columns].sum().reset_index()
    else:
        if not isinstance(aggregations, dict):
            raise HTTPException(status_code=400, detail="args.aggregations must be an object")
        summary = dataframe.groupby(group_by, dropna=False).agg(aggregations).reset_index()
        summary.columns = [
            "_".join(str(part) for part in column if part)
            if isinstance(column, tuple)
            else str(column)
            for column in summary.columns
        ]

    return summary


def _json_records(dataframe: pd.DataFrame) -> list[dict[str, Any]]:
    return dataframe.where(pd.notnull(dataframe), None).to_dict(orient="records")


def _write_dataframe_to_worksheet(dataframe: pd.DataFrame, worksheet: Any) -> None:
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        worksheet.cell(row=1, column=column_index, value=str(column_name))

    for row_index, row_values in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=_excel_cell_value(value))


def _excel_cell_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except ValueError:
            return value
    return value


def _command_category(command: ExcelCommandName) -> ExcelCommandCategory:
    return "basic" if command in BASIC_EXCEL_COMMANDS else "advanced"


def _finish_openpyxl_response(
    request: ExcelExecuteRequest,
    source_path: Path,
    output_path: Path,
    workbook: Any,
    sheet: str | None,
    *,
    rows_affected: int = 0,
    cells_affected: int = 0,
    summary: str,
    data: list[dict[str, Any]] | None = None,
) -> ExcelExecuteResponse:
    workbook_base64 = _save_openpyxl_workbook(workbook, output_path, request.save_to_disk)
    return _response(
        request,
        source_path,
        output_path,
        sheet,
        rows_affected=rows_affected,
        cells_affected=cells_affected,
        summary=summary,
        data=data,
        workbook_base64=workbook_base64,
        saved_to_disk=request.save_to_disk,
    )


def _save_openpyxl_workbook(workbook: Any, output_path: Path, save_to_disk: bool) -> str | None:
    try:
        if save_to_disk:
            workbook.save(output_path)
            return None

        buffer = BytesIO()
        workbook.save(buffer)
        return _encode_bytes(buffer.getvalue())
    finally:
        workbook.close()


def _encode_bytes(content: bytes) -> str:
    return b64encode(content).decode("ascii")


def _response(
    request: ExcelExecuteRequest,
    source_path: Path,
    output_path: Path,
    sheet: str | None,
    *,
    rows_affected: int = 0,
    cells_affected: int = 0,
    summary: str,
    data: list[dict[str, Any]] | None = None,
    workbook_base64: str | None = None,
    saved_to_disk: bool = False,
) -> ExcelExecuteResponse:
    return ExcelExecuteResponse(
        command=request.command,
        category=_command_category(request.command),
        file_path=str(source_path),
        output_path=str(output_path),
        workbook_base64=workbook_base64,
        saved_to_disk=saved_to_disk,
        sheet=sheet,
        rows_affected=rows_affected,
        cells_affected=cells_affected,
        summary=summary,
        data=data,
    )
